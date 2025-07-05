from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# Excel file to store login records
EXCEL_FILE = 'student_logins.xlsx'

# Create workbook and sheet if not exists
def create_excel_file_if_not_exists():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "LoginRecords"
        ws.append(["Student Name", "Login Time"])
        wb.save(EXCEL_FILE)

# Append login record
def log_student_login(student_name):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["LoginRecords"]
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([student_name, current_time])
    wb.save(EXCEL_FILE)
    print(f"Login recorded for {student_name} at {current_time}")

# Simulate login
def student_login():
    student_name = input("Enter your name to log in: ").strip()
    if student_name:
        log_student_login(student_name)
    else:
        print("Name cannot be empty.")

# Main
if __name__ == "__main__":
    create_excel_file_if_not_exists()
    student_login()
